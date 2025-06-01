import os
import re
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Error: openpyxl module not found!")
    print("Please install it: pip install openpyxl")
    exit()


class MedicalCampDataEntry:
    def __init__(self, filename="medical_camp_data.xlsx"):
        self.filename = filename
        self.columns = [
            "Name",
            "Age",
            "Gender",
            "Phone",
            "Blood Pressure",
            "Blood Group",
            "Blood Sugar",
            "Weight",
            "Height (cm)",
            "Height (ft)",
            "BMI",
            "Health Comments",
            "Date Added",
        ]
        self.data = []
        self.load_existing_data()

    def load_existing_data(self):
        """Load existing data from Excel file if it exists"""
        if os.path.exists(self.filename):
            try:
                wb = openpyxl.load_workbook(self.filename)
                ws = wb.active

                # Read header row to get column mapping
                headers = []
                for col in range(1, ws.max_column + 1):
                    header_cell = ws.cell(row=1, column=col)
                    if header_cell.value:
                        headers.append(str(header_cell.value))
                    else:
                        break

                # Read data rows
                for row in range(2, ws.max_row + 1):
                    patient = {}
                    has_data = False

                    for col, header in enumerate(headers, 1):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value is not None:
                            patient[header] = str(cell_value)
                            has_data = True
                        else:
                            patient[header] = ""

                    # Only add row if it has some data
                    if has_data and patient.get("Name", "").strip():
                        self.data.append(patient)

                wb.close()
                print(
                    f"✅ Loaded {len(self.data)} existing records from {self.filename}"
                )

            except Exception as e:
                print(f"⚠️  Could not load existing Excel file: {e}")
                print("Creating new data set...")
                self.data = []
        else:
            print(f"📝 Creating new Excel file: {self.filename}")

    def calculate_bmi(self, weight, height_cm):
        """Calculate BMI from weight (kg) and height (cm)"""
        try:
            weight_kg = float(weight)
            height_cm_val = float(height_cm)
            height_m = height_cm_val / 100  # Convert cm to meters
            bmi = weight_kg / (height_m**2)
            return round(bmi, 1)
        except (ValueError, ZeroDivisionError):
            return ""

    def categorize_bmi(self, bmi):
        """Categorize BMI with detailed classification"""
        try:
            bmi_val = float(bmi)
            if bmi_val < 16:
                return "Severely Underweight"
            elif 16 <= bmi_val < 18.5:
                return "Underweight"
            elif 18.5 <= bmi_val < 25:
                return "Normal"
            elif 25 <= bmi_val < 30:
                return "Overweight"
            elif 30 <= bmi_val < 35:
                return "Obese Class I"
            elif 35 <= bmi_val < 40:
                return "Obese Class II"
            else:
                return "Obese Class III"
        except ValueError:
            return ""

    def feet_to_cm(self, feet_input):
        """Convert feet'inches to centimeters"""
        try:
            # Handle formats like "5'6", "5.5", "5 6", "5'6\"", "5 feet 6 inches"
            feet_input = feet_input.strip().lower()

            # Remove common words
            feet_input = re.sub(r"\b(feet|foot|ft|inches|inch|in)\b", "", feet_input)
            feet_input = re.sub(r'["\']', " ", feet_input)  # Replace quotes with spaces

            # Extract numbers
            numbers = re.findall(r"\d+\.?\d*", feet_input)

            if len(numbers) == 1:
                # Single number - could be decimal feet or just feet
                num = float(numbers[0])
                if num > 10:  # Assume it's already in cm if > 10
                    return num
                else:  # Assume it's feet
                    return num * 30.48
            elif len(numbers) == 2:
                # Two numbers - feet and inches
                feet = float(numbers[0])
                inches = float(numbers[1])
                total_inches = (feet * 12) + inches
                return total_inches * 2.54
            else:
                return None
        except:
            return None

    def cm_to_feet(self, cm):
        """Convert centimeters to feet'inches format"""
        try:
            cm_val = float(cm)
            total_inches = cm_val / 2.54
            feet = int(total_inches // 12)
            inches = round(total_inches % 12)
            return f"{feet}'{inches}\""
        except:
            return ""

    def analyze_blood_pressure(self, bp_str):

        if not bp_str or "/" not in bp_str:
            return "❌ Invalid format. Use systolic/diastolic (e.g., 120/80)"

        try:
            systolic, diastolic = bp_str.split("/")
            sys_val = int(systolic.strip())
            dia_val = int(diastolic.strip())

            if sys_val >= 180 or dia_val >= 120:
                return "🚨 HYPERTENSIVE CRISIS - Emergency medical care needed"
            elif (140 <= sys_val <= 179) or (90 <= dia_val <= 119):
                return "🚨 HIGH BP - Immediate medical attention"
            elif (130 <= sys_val <= 139) or (80 <= dia_val <= 89):
                return "⚠️ BP Slightly higher than normal. Consider consulting doctor."
            elif (120 <= sys_val <= 129) and dia_val < 80:
                return "⚡ ELEVATED BP - Lifestyle changes recommended"
            elif sys_val < 120 and dia_val < 80:
                return "✅ BP is normal."
            elif sys_val < 90 or dia_val < 60:
                return "⚠️ BP is lower than normal."
            else:
                return "📋 BP recorded"

        except ValueError:
            return "❌ Invalid input. Please provide numeric values like '120/80'"

    def analyze_blood_sugar(self, sugar_str, test_type="random"):
        """Analyze blood sugar and return health comment (mmol/L format)"""
        if not sugar_str:
            return ""

        try:
            sugar_val = float(sugar_str)

            # Blood sugar interpretation in mmol/L
            if sugar_val < 3.9:
                return "⚠️ LOW SUGAR - Hypoglycemia risk"
            elif 3.9 <= sugar_val < 5.6:
                return "✅ NORMAL SUGAR - Good glucose level"
            elif 5.6 <= sugar_val < 7.8:
                return "⚡ BORDERLINE - Monitor glucose levels"
            elif 7.8 <= sugar_val < 11.1:
                return "⚠️ HIGH SUGAR - Pre-diabetic range"
            elif sugar_val >= 11.1:
                return "🚨 VERY HIGH SUGAR - Diabetic range, see doctor"
            else:
                return "📋 Sugar level recorded"

        except ValueError:
            return ""

    def analyze_bmi_health(self, bmi_str):
        """Analyze BMI and return health recommendations"""
        if not bmi_str:
            return ""

        try:
            bmi_val = float(bmi_str)

            if bmi_val < 16:
                return "🚨 SEVERELY UNDERWEIGHT - Nutritional support needed"
            elif 16 <= bmi_val < 18.5:
                return "⚠️ UNDERWEIGHT - Increase caloric intake"
            elif 18.5 <= bmi_val < 25:
                return "✅ HEALTHY WEIGHT - Maintain current lifestyle"
            elif 25 <= bmi_val < 30:
                return "⚡ OVERWEIGHT - Diet and exercise recommended"
            elif 30 <= bmi_val < 35:
                return "⚠️ OBESE - Medical consultation advised"
            elif 35 <= bmi_val < 40:
                return "🚨 SEVERELY OBESE - Immediate medical attention"
            else:
                return "🚨 MORBIDLY OBESE - Urgent medical intervention"

        except ValueError:
            return ""

    def validate_phone(self, phone):
        """Validate and format phone number"""
        if not phone:
            return ""

        # Remove all non-digits
        digits = re.sub(r"\D", "", phone)

        # Check for valid BD mobile number patterns
        if len(digits) == 11 and digits.startswith("01"):
            return f"+88{digits}"
        elif len(digits) == 13 and digits.startswith("880"):
            return f"+{digits}"
        elif len(digits) == 10:
            return f"+8801{digits}"
        else:
            return phone  # Return as-is if doesn't match patterns

    def add_patient(self):
        """Add a new patient record with enhanced validation"""
        print("\n" + "=" * 50)
        print("           ADD NEW PATIENT")
        print("=" * 50)
        patient = {}

        # Basic information
        patient["Name"] = input("👤 Patient Name: ").strip().title()
        if not patient["Name"]:
            print("❌ Name cannot be empty!")
            return

        # Age validation
        while True:
            try:
                age_input = input("🎂 Age: ").strip()
                age_val = int(age_input)
                if 0 <= age_val <= 120:
                    patient["Age"] = str(age_val)
                    break
                else:
                    print("❌ Please enter a valid age (0-120)")
            except ValueError:
                print("❌ Please enter a valid number for age")

        # Gender
        while True:
            gender = input("⚥ Gender (M/F/Male/Female): ").strip().upper()
            if gender in ["M", "F", "MALE", "FEMALE"]:
                patient["Gender"] = "Male" if gender in ["M", "MALE"] else "Female"
                break
            else:
                print("❌ Please enter M/F or Male/Female")

        # Phone number
        phone_input = input("📱 Phone Number (optional): ").strip()
        patient["Phone"] = self.validate_phone(phone_input)

        # Blood Pressure with enhanced validation
        print("\n🩺 VITAL SIGNS")
        print("-" * 30)
        while True:
            bp_input = input(
                "❤️  Blood Pressure (e.g., 120/80) [Enter to skip]: "
            ).strip()
            if not bp_input:
                patient["Blood Pressure"] = ""
                break

            if "/" in bp_input:
                try:
                    sys_str, dia_str = bp_input.split("/")
                    sys_val = int(sys_str.strip())
                    dia_val = int(dia_str.strip())

                    if (
                        50 <= sys_val <= 250
                        and 30 <= dia_val <= 150
                        and sys_val > dia_val
                    ):
                        patient["Blood Pressure"] = f"{sys_val}/{dia_val}"
                        break
                    else:
                        print(
                            "❌ Invalid BP range. Systolic: 50-250, Diastolic: 30-150"
                        )
                except ValueError:
                    print("❌ Invalid format. Use format like 120/80")
            else:
                print("❌ Please use format like 120/80")

        # Blood Group
        print("\n🩸 Blood Group Options: A+, A-, B+, B-, AB+, AB-, O+, O-")
        while True:
            blood_group = input("🩸 Blood Group [Enter to skip]: ").strip().upper()
            valid_groups = ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]
            if blood_group in valid_groups or blood_group == "":
                patient["Blood Group"] = blood_group
                break
            else:
                print("❌ Please enter a valid blood group or press Enter to skip")

        while True:
            sugar_input = input("🍯 Blood Sugar (mmol/L) [Enter to skip]: ").strip()
            if sugar_input == "":
                patient["Blood Sugar"] = ""
                break
            try:
                sugar_val = float(sugar_input)
                if 1.0 <= sugar_val <= 44.4:  # Reasonable range for mmol/L
                    patient["Blood Sugar"] = (
                        f"{sugar_val:.1f}"
                        if sugar_val != int(sugar_val)
                        else str(int(sugar_val))
                    )
                    break
                else:
                    print("❌ Please enter blood sugar between 1.0-44.4 mmol/L")
            except ValueError:
                print("❌ Please enter a valid number for blood sugar")

        # Weight
        print("\n📏 PHYSICAL MEASUREMENTS")
        print("-" * 30)
        while True:
            weight_input = input("⚖️  Weight (kg) [Enter to skip]: ").strip()
            if weight_input == "":
                patient["Weight"] = ""
                break
            try:
                weight_val = float(weight_input)
                if 1 <= weight_val <= 500:  # Reasonable range
                    patient["Weight"] = (
                        f"{weight_val:.1f}"
                        if weight_val != int(weight_val)
                        else str(int(weight_val))
                    )
                    break
                else:
                    print("❌ Please enter weight between 1-500 kg")
            except ValueError:
                print("❌ Please enter a valid number for weight")

        # Height with feet/cm options
        print("📐 Height can be entered in:")
        print("   - Centimeters (e.g., 170)")
        print("   - Feet'Inches (e.g., 5'6 or 5.5)")

        while True:
            height_input = input("📐 Height [Enter to skip]: ").strip()
            if height_input == "":
                patient["Height (cm)"] = ""
                patient["Height (ft)"] = ""
                break

            # Try to parse as feet first
            height_cm = self.feet_to_cm(height_input)
            if height_cm and 30 <= height_cm <= 300:
                patient["Height (cm)"] = (
                    f"{height_cm:.1f}"
                    if height_cm != int(height_cm)
                    else str(int(height_cm))
                )
                patient["Height (ft)"] = self.cm_to_feet(height_cm)
                break

            # Try as direct cm input
            try:
                height_val = float(height_input)
                if 30 <= height_val <= 300:
                    patient["Height (cm)"] = (
                        f"{height_val:.1f}"
                        if height_val != int(height_val)
                        else str(int(height_val))
                    )
                    patient["Height (ft)"] = self.cm_to_feet(height_val)
                    break
                else:
                    print("❌ Please enter height between 30-300 cm or 1'-10'")
            except ValueError:
                print("❌ Invalid height format. Try: 170 or 5'6")

        # Calculate BMI
        if patient["Weight"] and patient["Height (cm)"]:
            bmi = self.calculate_bmi(patient["Weight"], patient["Height (cm)"])
            patient["BMI"] = str(bmi) if bmi else ""
        else:
            patient["BMI"] = ""

        # Generate health comments
        health_comments = []

        # Analyze blood pressure
        bp_comment = self.analyze_blood_pressure(patient["Blood Pressure"])
        if bp_comment:
            health_comments.append(bp_comment)

        # Analyze blood sugar
        sugar_comment = self.analyze_blood_sugar(patient["Blood Sugar"])
        if sugar_comment:
            health_comments.append(sugar_comment)

        # Analyze BMI
        bmi_comment = self.analyze_bmi_health(patient["BMI"])
        if bmi_comment:
            health_comments.append(bmi_comment)

        # Additional health observations based on age
        age_val = int(patient["Age"])
        if age_val >= 60:
            health_comments.append("👴 SENIOR - Regular health checkups recommended")
        elif age_val <= 2:
            health_comments.append("👶 INFANT - Pediatric care recommended")
        elif age_val <= 12:
            health_comments.append("🧒 CHILD - Growth monitoring important")
        elif age_val <= 19:
            health_comments.append("👦 ADOLESCENT - Developmental checkups advised")

        patient["Health Comments"] = " | ".join(health_comments)

        # Add timestamp
        patient["Date Added"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        self.data.append(patient)

        # Display comprehensive patient summary
        print("\n" + "=" * 60)
        print("               PATIENT SUMMARY")
        print("=" * 60)
        print(f"👤 Name: {patient['Name']}")
        print(f"🎂 Age: {patient['Age']} years ({patient['Gender']})")
        if patient["Phone"]:
            print(f"📱 Phone: {patient['Phone']}")

        print(f"\n🩺 VITAL SIGNS:")
        if patient["Blood Pressure"]:
            bp_analysis = self.analyze_blood_pressure(patient["Blood Pressure"])
            print(f"   ❤️  Blood Pressure: {patient['Blood Pressure']} mmHg")
            if bp_analysis:
                print(f"      → {bp_analysis}")

        if patient["Blood Group"]:
            print(f"   🩸 Blood Group: {patient['Blood Group']}")

        if patient["Blood Sugar"]:
            sugar_analysis = self.analyze_blood_sugar(patient["Blood Sugar"])
            print(f"   🍯 Blood Sugar: {patient['Blood Sugar']} mmol/L")
            if sugar_analysis:
                print(f"      → {sugar_analysis}")

        print(f"\n📏 PHYSICAL MEASUREMENTS:")
        if patient["Weight"]:
            print(f"   ⚖️  Weight: {patient['Weight']} kg")
        if patient["Height (cm)"]:
            print(
                f"   📐 Height: {patient['Height (cm)']} cm ({patient['Height (ft)']})"
            )
        if patient["BMI"]:
            bmi_category = self.categorize_bmi(patient["BMI"])
            bmi_analysis = self.analyze_bmi_health(patient["BMI"])
            print(f"   📊 BMI: {patient['BMI']} ({bmi_category})")
            if bmi_analysis:
                print(f"      → {bmi_analysis}")

        if patient["Health Comments"]:
            print(f"\n💡 HEALTH INSIGHTS:")
            for comment in patient["Health Comments"].split(" | "):
                print(f"   • {comment}")

        print(f"\n📅 Recorded: {patient['Date Added']}")
        print("=" * 60)
        print("✅ Patient record added successfully!")

        # Auto-save to Excel after adding
        self.save_to_excel()

    def view_patients(self):
        """View all patients with enhanced display"""
        if not self.data:
            print("\n📋 No patients recorded yet.")
            return

        print(f"\n" + "=" * 80)
        print(f"           PATIENT RECORDS ({len(self.data)} total)")
        print("=" * 80)

        for i, patient in enumerate(self.data, 1):
            print(
                f"\n{i}. 👤 {patient['Name']} (Age: {patient['Age']}, {patient['Gender']})"
            )
            if patient.get("Phone", ""):
                print(f"   📱 {patient['Phone']}")

            # Display measurements in organized format
            measurements = []
            if patient.get("Blood Pressure", ""):
                measurements.append(f"BP: {patient['Blood Pressure']}")
            if patient.get("Blood Group", ""):
                measurements.append(f"Blood: {patient['Blood Group']}")
            if patient.get("Blood Sugar", ""):
                measurements.append(f"Sugar: {patient['Blood Sugar']} mmol/L")

            physical = []
            if patient.get("Weight", ""):
                physical.append(f"{patient['Weight']}kg")
            if patient.get("Height (cm)", ""):
                physical.append(f"{patient['Height (ft)']}")
            if patient.get("BMI", ""):
                bmi_cat = self.categorize_bmi(patient["BMI"])
                physical.append(f"BMI: {patient['BMI']} ({bmi_cat})")

            if measurements:
                print(f"   🩺 {' | '.join(measurements)}")
            if physical:
                print(f"   📏 {' | '.join(physical)}")

            if patient.get("Health Comments", ""):
                print(f"   💡 Health Notes:")
                for comment in patient["Health Comments"].split(" | "):
                    print(f"      • {comment}")

            print(f"   📅 {patient.get('Date Added', 'N/A')}")
            print("-" * 80)

    def search_patient(self):
        """Enhanced patient search with multiple criteria"""
        if not self.data:
            print("\n📋 No patients recorded yet.")
            return

        print("\n🔍 SEARCH OPTIONS:")
        print("1. Search by Name")
        print("2. Search by Blood Group")
        print("3. Search by Age Range")
        print("4. Search by Health Condition")

        choice = input("\nSelect search type (1-4): ").strip()

        if choice == "1":
            self._search_by_name()
        elif choice == "2":
            self._search_by_blood_group()
        elif choice == "3":
            self._search_by_age()
        elif choice == "4":
            self._search_by_health()
        else:
            print("❌ Invalid choice")

    def _search_by_name(self):
        """Search patients by name"""
        search_name = input("\n👤 Enter patient name to search: ").strip().lower()
        found_patients = []

        for i, patient in enumerate(self.data):
            if search_name in patient["Name"].lower():
                found_patients.append((i, patient))

        self._display_search_results(found_patients, f"name containing '{search_name}'")

    def _search_by_blood_group(self):
        """Search patients by blood group"""
        blood_group = input("\n🩸 Enter blood group (e.g., A+, O-): ").strip().upper()
        found_patients = []

        for i, patient in enumerate(self.data):
            if patient.get("Blood Group", "").upper() == blood_group:
                found_patients.append((i, patient))

        self._display_search_results(found_patients, f"blood group {blood_group}")

    def _search_by_age(self):
        """Search patients by age range"""
        try:
            min_age = int(input("\n🎂 Minimum age: ").strip())
            max_age = int(input("🎂 Maximum age: ").strip())

            found_patients = []
            for i, patient in enumerate(self.data):
                age = int(patient.get("Age", 0))
                if min_age <= age <= max_age:
                    found_patients.append((i, patient))

            self._display_search_results(
                found_patients, f"age between {min_age}-{max_age}"
            )
        except ValueError:
            print("❌ Invalid age format")

    def _search_by_health(self):
        """Search patients by health conditions"""
        print("\n💡 Common health conditions:")
        print("- high bp, low bp, normal bp")
        print("- high sugar, low sugar, normal sugar")
        print("- underweight, overweight, obese")

        condition = input("\n🔍 Enter condition to search: ").strip().lower()
        found_patients = []

        for i, patient in enumerate(self.data):
            health_comments = patient.get("Health Comments", "").lower()
            if condition in health_comments:
                found_patients.append((i, patient))

        self._display_search_results(found_patients, f"condition '{condition}'")

    def _display_search_results(self, found_patients, search_criteria):
        """Display search results"""
        if found_patients:
            print(
                f"\n✅ Found {len(found_patients)} patient(s) with {search_criteria}:"
            )
            print("=" * 60)

            for i, patient in found_patients:
                print(
                    f"\n👤 {patient['Name']} (Age: {patient['Age']}, {patient['Gender']})"
                )

                if patient.get("Phone", ""):
                    print(f"📱 {patient['Phone']}")
                if patient.get("Blood Pressure", ""):
                    print(f"❤️  BP: {patient['Blood Pressure']} mmHg")
                if patient.get("Blood Group", ""):
                    print(f"🩸 Blood: {patient['Blood Group']}")
                if patient.get("Blood Sugar", ""):
                    print(f"🍯 Sugar: {patient['Blood Sugar']} mmol/L")
                if patient.get("Weight", "") and patient.get("Height (cm)", ""):
                    print(
                        f"📏 Physical: {patient['Weight']} kg, {patient['Height (ft)']} ({patient['Height (cm)']} cm)"
                    )
                    if patient.get("BMI", ""):
                        bmi_cat = self.categorize_bmi(patient["BMI"])
                        print(f"📊 BMI: {patient['BMI']} ({bmi_cat})")

                if patient.get("Health Comments", ""):
                    print(f"💡 Health Notes:")
                    for comment in patient["Health Comments"].split(" | "):
                        print(f"   • {comment}")

                print(f"📅 {patient.get('Date Added', 'N/A')}")
                print("-" * 60)
        else:
            print(f"\n❌ No patients found with {search_criteria}")

    def save_to_excel(self):
        """Save data to Excel file with professional formatting"""
        if not self.data:
            print("📋 No data to save.")
            return

        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Medical Camp Data"

            # Write headers with formatting
            headers = self.columns
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")

            # Write data
            for row, patient in enumerate(self.data, 2):
                for col, header in enumerate(headers, 1):
                    value = patient.get(header, "")
                    ws.cell(row=row, column=col, value=value)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 3, 40)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add borders and alternating row colors
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Apply formatting to all cells
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Alternate row colors (light blue for even rows)
            light_fill = PatternFill(
                start_color="F0F8FF", end_color="F0F8FF", fill_type="solid"
            )
            for row in range(3, len(self.data) + 2, 2):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = light_fill

            # Create summary sheet
            summary_ws = wb.create_sheet("Summary Report")
            self._create_summary_sheet(summary_ws)

            # Create health alerts sheet
            alerts_ws = wb.create_sheet("Health Alerts")
            self._create_health_alerts_sheet(alerts_ws)

            # Save the workbook
            wb.save(self.filename)

            print(f"💾 Data saved to Excel: {self.filename}")
            print(f"📊 Total records: {len(self.data)}")
            print(f"📄 Sheets created: Patient Data, Summary Report, Health Alerts")

        except Exception as e:
            print(f"❌ Error saving Excel file: {e}")

    def _create_health_alerts_sheet(self, ws):
        """Create health alerts sheet for patients needing attention"""
        # Title
        ws["A1"] = "HEALTH ALERTS & PRIORITY PATIENTS"
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color="DC143C", end_color="DC143C", fill_type="solid"
        )
        ws.merge_cells("A1:F1")

        row = 3

        # High priority alerts
        high_priority = []
        medium_priority = []

        for patient in self.data:
            alerts = []
            priority = "LOW"

            # Check blood pressure
            bp_comment = self.analyze_blood_pressure(patient.get("Blood Pressure", ""))
            if "CRISIS" in bp_comment or "STAGE 2" in bp_comment:
                alerts.append(bp_comment)
                priority = "HIGH"
            elif "STAGE 1" in bp_comment or "LOW BP" in bp_comment:
                alerts.append(bp_comment)
                if priority != "HIGH":
                    priority = "MEDIUM"

            # Check blood sugar
            sugar_comment = self.analyze_blood_sugar(patient.get("Blood Sugar", ""))
            if "VERY HIGH" in sugar_comment or "LOW SUGAR" in sugar_comment:
                alerts.append(sugar_comment)
                priority = "HIGH"
            elif "HIGH SUGAR" in sugar_comment:
                alerts.append(sugar_comment)
                if priority != "HIGH":
                    priority = "MEDIUM"

            # Check BMI
            bmi_comment = self.analyze_bmi_health(patient.get("BMI", ""))
            if "SEVERELY" in bmi_comment or "MORBIDLY" in bmi_comment:
                alerts.append(bmi_comment)
                priority = "HIGH"
            elif "OBESE" in bmi_comment:
                alerts.append(bmi_comment)
                if priority != "HIGH":
                    priority = "MEDIUM"

            if alerts:
                patient_alert = {
                    "name": patient["Name"],
                    "age": patient["Age"],
                    "gender": patient["Gender"],
                    "phone": patient.get("Phone", ""),
                    "alerts": alerts,
                    "priority": priority,
                }

                if priority == "HIGH":
                    high_priority.append(patient_alert)
                else:
                    medium_priority.append(patient_alert)

        # High Priority Section
        if high_priority:
            ws[f"A{row}"] = "🚨 HIGH PRIORITY ALERTS"
            ws[f"A{row}"].font = Font(bold=True, size=14, color="FFFFFF")
            ws[f"A{row}"].fill = PatternFill(
                start_color="DC143C", end_color="DC143C", fill_type="solid"
            )
            ws.merge_cells(f"A{row}:F{row}")
            row += 2

            # Headers
            headers = [
                "Patient Name",
                "Age",
                "Gender",
                "Phone",
                "Health Alerts",
                "Action Required",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="8B0000", end_color="8B0000", fill_type="solid"
                )
            row += 1

            for alert in high_priority:
                ws.cell(row=row, column=1, value=alert["name"])
                ws.cell(row=row, column=2, value=alert["age"])
                ws.cell(row=row, column=3, value=alert["gender"])
                ws.cell(row=row, column=4, value=alert["phone"])
                ws.cell(row=row, column=5, value=" | ".join(alert["alerts"]))
                ws.cell(row=row, column=6, value="Immediate medical attention required")

                # Color code the row
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFE4E1", end_color="FFE4E1", fill_type="solid"
                    )

                row += 1
            row += 1

        # Medium Priority Section
        if medium_priority:
            ws[f"A{row}"] = "⚠️ MEDIUM PRIORITY ALERTS"
            ws[f"A{row}"].font = Font(bold=True, size=14, color="FFFFFF")
            ws[f"A{row}"].fill = PatternFill(
                start_color="FF8C00", end_color="FF8C00", fill_type="solid"
            )
            ws.merge_cells(f"A{row}:F{row}")
            row += 2

            # Headers
            headers = [
                "Patient Name",
                "Age",
                "Gender",
                "Phone",
                "Health Alerts",
                "Recommendation",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="FF6347", end_color="FF6347", fill_type="solid"
                )
            row += 1

            for alert in medium_priority:
                ws.cell(row=row, column=1, value=alert["name"])
                ws.cell(row=row, column=2, value=alert["age"])
                ws.cell(row=row, column=3, value=alert["gender"])
                ws.cell(row=row, column=4, value=alert["phone"])
                ws.cell(row=row, column=5, value=" | ".join(alert["alerts"]))
                ws.cell(row=row, column=6, value="Follow-up recommended")

                # Color code the row
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFF8DC", end_color="FFF8DC", fill_type="solid"
                    )

                row += 1

        # Summary statistics
        row += 2
        ws[f"A{row}"] = "ALERT SUMMARY"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        ws[f"A{row}"] = f"High Priority Patients: {len(high_priority)}"
        ws[f"A{row}"].font = Font(color="DC143C")
        row += 1
        ws[f"A{row}"] = f"Medium Priority Patients: {len(medium_priority)}"
        ws[f"A{row}"].font = Font(color="FF8C00")
        row += 1
        ws[f"A{row}"] = (
            f"Total Patients Needing Follow-up: {len(high_priority) + len(medium_priority)}"
        )
        ws[f"A{row}"].font = Font(bold=True)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 3, 35)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_summary_sheet(self, ws):
        """Create a comprehensive summary sheet in the Excel workbook"""
        # Title
        ws["A1"] = "HASHIMUKH MEDICAL CAMP COMPREHENSIVE REPORT"
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        ws.merge_cells("A1:D1")

        row = 3

        # Basic statistics
        ws[f"A{row}"] = "BASIC STATISTICS"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        ws[f"A{row}"] = f"Total Patients Registered:"
        ws[f"B{row}"] = len(self.data)
        row += 1

        # Gender statistics
        male_count = sum(1 for p in self.data if p.get("Gender", "").lower() == "male")
        female_count = len(self.data) - male_count

        ws[f"A{row}"] = f"Male Patients:"
        ws[f"B{row}"] = f"{male_count} ({male_count/len(self.data)*100:.1f}%)"
        row += 1
        ws[f"A{row}"] = f"Female Patients:"
        ws[f"B{row}"] = f"{female_count} ({female_count/len(self.data)*100:.1f}%)"
        row += 1

        # Age statistics
        ages = [int(p["Age"]) for p in self.data if str(p["Age"]).isdigit()]
        if ages:
            ws[f"A{row}"] = f"Age Range:"
            ws[f"B{row}"] = f"{min(ages)} - {max(ages)} years"
            row += 1
            ws[f"A{row}"] = f"Average Age:"
            ws[f"B{row}"] = f"{sum(ages)/len(ages):.1f} years"
            row += 1
            ws[f"A{row}"] = f"Median Age:"
            ws[f"B{row}"] = f"{sorted(ages)[len(ages)//2]:.1f} years"
            row += 2

        # Blood Group Distribution
        ws[f"A{row}"] = "BLOOD GROUP DISTRIBUTION"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        blood_groups = {}
        tested_for_blood_group = 0
        for p in self.data:
            bg = p.get("Blood Group", "").strip()
            if bg:
                blood_groups[bg] = blood_groups.get(bg, 0) + 1
                tested_for_blood_group += 1

        if blood_groups:
            ws[f"A{row}"] = "Blood Group"
            ws[f"B{row}"] = "Count"
            ws[f"C{row}"] = "Percentage"

            # Header formatting
            for col in ["A", "B", "C"]:
                ws[f"{col}{row}"].font = Font(bold=True)
                ws[f"{col}{row}"].fill = PatternFill(
                    start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
                )

            row += 1

            for bg in sorted(blood_groups.keys()):
                count = blood_groups[bg]
                percentage = (count / tested_for_blood_group) * 100
                ws[f"A{row}"] = bg
                ws[f"B{row}"] = count
                ws[f"C{row}"] = f"{percentage:.1f}%"
                row += 1

        ws[f"A{row}"] = f"Not Tested for Blood Group:"
        ws[f"B{row}"] = len(self.data) - tested_for_blood_group
        row += 2

        # BMI Statistics
        ws[f"A{row}"] = "BMI STATISTICS"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        bmi_categories = {
            "Severely Underweight": 0,
            "Underweight": 0,
            "Normal": 0,
            "Overweight": 0,
            "Obese Class I": 0,
            "Obese Class II": 0,
            "Obese Class III": 0,
        }
        measured_count = 0
        bmi_values = []

        for p in self.data:
            if p.get("BMI", ""):
                measured_count += 1
                bmi_val = float(p["BMI"])
                bmi_values.append(bmi_val)
                bmi_cat = self.categorize_bmi(p["BMI"])
                if bmi_cat in bmi_categories:
                    bmi_categories[bmi_cat] += 1

        if measured_count > 0:
            ws[f"A{row}"] = "BMI Category"
            ws[f"B{row}"] = "Count"
            ws[f"C{row}"] = "Percentage"

            # Header formatting
            for col in ["A", "B", "C"]:
                ws[f"{col}{row}"].font = Font(bold=True)
                ws[f"{col}{row}"].fill = PatternFill(
                    start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
                )

            row += 1

            for category, count in bmi_categories.items():
                if count > 0:
                    percentage = (count / measured_count) * 100
                    ws[f"A{row}"] = category
                    ws[f"B{row}"] = count
                    ws[f"C{row}"] = f"{percentage:.1f}%"
                    row += 1

            # BMI statistics
            row += 1
            ws[f"A{row}"] = f"Average BMI:"
            ws[f"B{row}"] = f"{sum(bmi_values)/len(bmi_values):.1f}"
            row += 1
            ws[f"A{row}"] = f"BMI Range:"
            ws[f"B{row}"] = f"{min(bmi_values):.1f} - {max(bmi_values):.1f}"

        ws[f"A{row+1}"] = f"Not Measured for BMI:"
        ws[f"B{row+1}"] = len(self.data) - measured_count
        row += 3

        # Health Screening Coverage
        ws[f"A{row}"] = "HEALTH SCREENING COVERAGE"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        sugar_tested = sum(1 for p in self.data if p.get("Blood Sugar", ""))
        bp_tested = sum(1 for p in self.data if p.get("Blood Pressure", ""))
        phone_provided = sum(1 for p in self.data if p.get("Phone", ""))

        tests = [
            ("Blood Group Testing", tested_for_blood_group),
            ("Blood Sugar Testing", sugar_tested),
            ("Blood Pressure Check", bp_tested),
            ("BMI Calculation", measured_count),
            ("Contact Information", phone_provided),
        ]

        ws[f"A{row}"] = "Screening Type"
        ws[f"B{row}"] = "Completed"
        ws[f"C{row}"] = "Coverage %"

        # Header formatting
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].font = Font(bold=True)
            ws[f"{col}{row}"].fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

        row += 1

        for test_name, count in tests:
            percentage = (count / len(self.data)) * 100
            ws[f"A{row}"] = test_name
            ws[f"B{row}"] = f"{count}/{len(self.data)}"
            ws[f"C{row}"] = f"{percentage:.1f}%"
            row += 1

        # Health Alert Summary
        row += 2
        ws[f"A{row}"] = "HEALTH ALERT SUMMARY"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="DC143C")
        row += 1

        high_bp = sum(
            1
            for p in self.data
            if "HIGH BP" in self.analyze_blood_pressure(p.get("Blood Pressure", ""))
        )
        high_sugar = sum(
            1
            for p in self.data
            if "HIGH SUGAR" in self.analyze_blood_sugar(p.get("Blood Sugar", ""))
        )
        obese = sum(
            1 for p in self.data if "OBESE" in self.analyze_bmi_health(p.get("BMI", ""))
        )

        ws[f"A{row}"] = f"Patients with High Blood Pressure:"
        ws[f"B{row}"] = high_bp
        row += 1
        ws[f"A{row}"] = f"Patients with High Blood Sugar:"
        ws[f"B{row}"] = high_sugar
        row += 1
        ws[f"A{row}"] = f"Patients with Obesity:"
        ws[f"B{row}"] = obese
        row += 1
        ws[f"A{row}"] = f"Total Patients Needing Follow-up:"
        ws[f"B{row}"] = high_bp + high_sugar + obese
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font(bold=True)

        # Auto-adjust column widths for summary sheet
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 3, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def print_summary_report(self):
        """Print a comprehensive summary report of the medical camp"""
        if not self.data:
            print("\n📋 No data available for report.")
            return

        print(f"\n" + "=" * 80)
        print(f"        HASHIMUKH MEDICAL CAMP COMPREHENSIVE REPORT")
        print("=" * 80)
        print(f"📅 Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"👥 Total Patients Registered: {len(self.data)}")

        # Gender statistics
        male_count = sum(1 for p in self.data if p.get("Gender", "").lower() == "male")
        female_count = len(self.data) - male_count
        print(f"\n👫 GENDER DISTRIBUTION:")
        print(f"   👨 Male: {male_count} ({male_count/len(self.data)*100:.1f}%)")
        print(f"   👩 Female: {female_count} ({female_count/len(self.data)*100:.1f}%)")

        # Age statistics
        ages = [int(p["Age"]) for p in self.data if str(p["Age"]).isdigit()]
        if ages:
            print(f"\n🎂 AGE STATISTICS:")
            print(f"   📊 Age Range: {min(ages)} - {max(ages)} years")
            print(f"   📈 Average Age: {sum(ages)/len(ages):.1f} years")
            print(f"   📉 Median Age: {sorted(ages)[len(ages)//2]} years")

        # Blood Group Distribution
        blood_groups = {}
        tested_for_blood_group = 0
        for p in self.data:
            bg = p.get("Blood Group", "").strip()
            if bg:
                blood_groups[bg] = blood_groups.get(bg, 0) + 1
                tested_for_blood_group += 1

        if blood_groups:
            print(f"\n🩸 BLOOD GROUP DISTRIBUTION ({tested_for_blood_group} tested):")
            for bg in sorted(blood_groups.keys()):
                count = blood_groups[bg]
                percentage = (count / tested_for_blood_group) * 100
                print(f"   {bg}: {count} patients ({percentage:.1f}%)")

        # BMI Statistics
        bmi_categories = {
            "Severely Underweight": 0,
            "Underweight": 0,
            "Normal": 0,
            "Overweight": 0,
            "Obese Class I": 0,
            "Obese Class II": 0,
            "Obese Class III": 0,
        }
        measured_count = 0
        bmi_values = []

        for p in self.data:
            if p.get("BMI", ""):
                measured_count += 1
                bmi_val = float(p["BMI"])
                bmi_values.append(bmi_val)
                bmi_cat = self.categorize_bmi(p["BMI"])
                if bmi_cat in bmi_categories:
                    bmi_categories[bmi_cat] += 1

        if measured_count > 0:
            print(f"\n📊 BMI DISTRIBUTION ({measured_count} measured):")
            for category, count in bmi_categories.items():
                if count > 0:
                    percentage = (count / measured_count) * 100
                    print(f"   {category}: {count} patients ({percentage:.1f}%)")
            print(f"   📈 Average BMI: {sum(bmi_values)/len(bmi_values):.1f}")
            print(f"   📉 BMI Range: {min(bmi_values):.1f} - {max(bmi_values):.1f}")

        # Health Screening Coverage
        sugar_tested = sum(1 for p in self.data if p.get("Blood Sugar", ""))
        bp_tested = sum(1 for p in self.data if p.get("Blood Pressure", ""))
        phone_provided = sum(1 for p in self.data if p.get("Phone", ""))

        print(f"\n🩺 HEALTH SCREENING COVERAGE:")
        print(
            f"   🩸 Blood Group: {tested_for_blood_group}/{len(self.data)} ({tested_for_blood_group/len(self.data)*100:.1f}%)"
        )
        print(
            f"   🍯 Blood Sugar: {sugar_tested}/{len(self.data)} ({sugar_tested/len(self.data)*100:.1f}%)"
        )
        print(
            f"   ❤️  Blood Pressure: {bp_tested}/{len(self.data)} ({bp_tested/len(self.data)*100:.1f}%)"
        )
        print(
            f"   📏 BMI Calculation: {measured_count}/{len(self.data)} ({measured_count/len(self.data)*100:.1f}%)"
        )
        print(
            f"   📱 Contact Info: {phone_provided}/{len(self.data)} ({phone_provided/len(self.data)*100:.1f}%)"
        )

        # Health Alert Summary
        high_bp = sum(
            1
            for p in self.data
            if "HIGH BP" in self.analyze_blood_pressure(p.get("Blood Pressure", ""))
        )
        high_sugar = sum(
            1
            for p in self.data
            if "HIGH SUGAR" in self.analyze_blood_sugar(p.get("Blood Sugar", ""))
        )
        obese = sum(
            1 for p in self.data if "OBESE" in self.analyze_bmi_health(p.get("BMI", ""))
        )
        crisis_bp = sum(
            1
            for p in self.data
            if "CRISIS" in self.analyze_blood_pressure(p.get("Blood Pressure", ""))
        )

        print(f"\n🚨 HEALTH ALERT SUMMARY:")
        print(f"   ❤️  High Blood Pressure: {high_bp} patients")
        print(f"   🍯 High Blood Sugar: {high_sugar} patients")
        print(f"   ⚖️  Obesity Cases: {obese} patients")
        if crisis_bp > 0:
            print(f"   🚨 CRITICAL - Hypertensive Crisis: {crisis_bp} patients")
        print(
            f"   📋 Total Requiring Follow-up: {high_bp + high_sugar + obese} patients"
        )

        print("=" * 80)

    def run(self):
        print("=" * 60)
        print("    🏥 HASHIMUKH MEDICAL CAMP DATA ENTRY SYSTEM")
        print("           Health Screening & Vital Signs")
        print("=" * 60)

        while True:
            print("\n📋 MAIN MENU")
            print("-" * 30)
            print("1. ➕ Add New Patient")
            print("2. 👀 View All Patients")
            print("3. 🔍 Search Patients")
            print("4. 💾 Save to Excel")
            print("5. 📊 Summary Report")
            print("6. 🚪 Exit")

            choice = input("\n🎯 Enter your choice (1-6): ").strip()

            if choice == "1":
                self.add_patient()
            elif choice == "2":
                self.view_patients()
            elif choice == "3":
                self.search_patient()
            elif choice == "4":
                self.save_to_excel()
            elif choice == "5":
                self.print_summary_report()
            elif choice == "6":
                print("\n💾 Saving data before exit...")
                self.save_to_excel()
                print("👋 Thank you for using Hashimukh Medical Camp System!")
                print("🏥 Stay healthy, stay safe!")
                break
            else:
                print("❌ Invalid choice. Please enter 1-6.")


# Run the program
if __name__ == "__main__":
    # Create and run the enhanced medical camp data entry system
    app = MedicalCampDataEntry()
    app.run()
